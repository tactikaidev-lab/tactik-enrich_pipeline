#!/usr/bin/env python3
"""
Lead Enrichment Pipeline — Tactik AI Dev
Monday task: Lead Scraping + Enrichment HANDOVER (item 2835598691)

v2 — headers verified against the real uploaded files (fresh_networking_scored.xlsx,
mazenod_alumni.xlsx, transcend_final.xlsx, vsbn_final.xlsx, whatsapp_contacts_bali.json).
Adds an Airtable push as the final step (CSV stays the source of truth; Airtable is
the last stage, not the storage layer — if the push fails you re-run it from the CSV,
you don't lose the enrichment work).

WHAT THIS DOES
  Reads each of the 5 raw lead sources, fills in missing email / company / LinkedIn /
  website / job title using Hunter.io + Apollo.io + People Data Labs, writes a clean
  CSV, then (optionally) pushes the rows into an Airtable base/table.

VERIFIED SOURCE STATS (checked directly against the uploaded files)
  Fresh Networking : 430 rows, 423 already have an email (98%) — near enrichment-complete,
                      just needs normalizing into the target schema. LOWEST priority.
  Transcend         : 514 rows, 283 have an email (55%), 173 have LinkedIn (34%)
  VSBN              : 447 rows, 235 have an email (53%), no LinkedIn column at all
  Mazenod Alumni    : 1,009 rows, 0 emails, 36 have LinkedIn (4%) — confirmed not
                      viably enrichable at scale, flagged not run by default
  Bali WhatsApp     : 138 contacts (39 named + 99 phone-only) — hardest, no company/site

REQUIRED ENV VARS (set at `docker run` time via --env-file, never baked into the image)
  HUNTER_API_KEY       (vsbn/transcend only — if unset, falls back to free site
                        scraping instead of erroring; see DESIGN NOTES below)
  APOLLO_API_KEY       (bali only)
  PDL_API_KEY          (bali only)
  AIRTABLE_API_KEY     (only needed if --airtable-push is used)
  AIRTABLE_BASE_ID     (only needed if --airtable-push is used)

USAGE
  python enrich_pipeline.py --source vsbn --input vsbn_final.xlsx --output vsbn_enriched.csv
  python enrich_pipeline.py --source transcend --input transcend_final.xlsx --output transcend_enriched.csv
  python enrich_pipeline.py --source fresh --input fresh_networking_scored.xlsx --output fresh_enriched.csv
  python enrich_pipeline.py --source mazenod --input mazenod_alumni.xlsx --output mazenod_enriched.csv
  python enrich_pipeline.py --source bali --input whatsapp_contacts_bali.json --output bali_enriched.csv

  --limit N                 process only the first N rows (validate on ~20 before a full run)
  --dry-run                 no paid API calls, just report what would happen
  --airtable-push           after writing the CSV, push rows into Airtable
  --airtable-table NAME     Airtable table name (required if --airtable-push)
  --smtp-verify             OPT-IN, vsbn/transcend only: when HUNTER_API_KEY is unset and the
                            free site-scrape finds nothing, also try pattern-guessed addresses
                            confirmed via real (no-send) SMTP RCPT TO. Off by default — see
                            DESIGN NOTES below for why.

DESIGN NOTES
  - Every API call is cached to a local sqlite db (enrich_cache.db), keyed by lookup
    key, so re-runs never pay twice for the same lookup. MX checks are cached too.
  - Hunter's email VERIFIER always runs on any Hunter-found email before it's accepted.
  - Known-blocked approaches (LinkedIn scraping, search-engine scraping, free
    reverse-phone lookup) are deliberately NOT implemented — already tried and
    failed from the server IP, per the brief.
  - Free fallback: if HUNTER_API_KEY isn't set, enrich_domain_based (vsbn/transcend)
    scrapes the LEAD'S OWN public site (homepage + /contact, /about style pages) for
    a listed email instead of calling Hunter. This is scoped to the lead's own
    domain only — it is NOT the LinkedIn/search-engine scraping ruled out above.
    Lower confidence than Hunter: no pattern-matching, and verification is an MX
    lookup (domain can receive mail) rather than Hunter's real mailbox check. Every
    row enriched this way gets a note in `_enrichment_notes` saying so, so it's
    flagged for review rather than silently trusted the same as a Hunter result.
    Once HUNTER_API_KEY is set, this fallback stops being used automatically.
    MEASURED: on a real 56-domain sample of VSBN leads (2026-09), this found 0
    emails — most modern small-business sites don't expose one in raw HTML
    anymore (contact forms / JS obfuscation instead). Treat it as a compliance
    fallback that costs nothing to leave on, not a real coverage strategy — a
    free Hunter.io key (25 finds + 50 verifications/month, no card) will
    outperform it because Hunter pattern-guesses candidates rather than only
    reading what's already published.
  - --smtp-verify (opt-in, off by default): the free equivalent of Hunter's
    pattern-guess-then-verify, when nothing's already published on the site.
    Generates likely addresses (first.last@domain, info@domain, etc.) and
    confirms each via a real RCPT TO against the lead's own mail server
    (never a third party), no message ever sent. Left off by default because
    it's a materially different footprint than the rest of this pipeline —
    an active connection to someone else's mail server, which some networks
    flag as probing, versus passive HTTP/DNS elsewhere. Also worth knowing:
    catch-all domains accept RCPT TO for anything, which shows up as a false
    "accepted" here — flagged in the note on every hit from this path.
  - Provider HTTP calls (Hunter/Apollo/PDL/Airtable) retry transient failures
    (timeouts, connection errors, 5xx) with exponential backoff via
    _request_with_retry before giving up — a single flaky response no longer
    needs a full re-run.
  - Per-row error isolation: if a row's enrichment call raises (bad data, an
    exhausted quota, a provider outage), that row is logged, flagged in
    `_enrichment_notes`, counted in the run summary, and the batch continues —
    it does not abort the rest of the rows. The CSV is written incrementally
    (flushed row-by-row) for the same reason: a hard kill leaves a usable
    partial file, not nothing.
  - Every run writes `<output>.summary.json` alongside the CSV — rows
    processed/enriched/flagged/errored, per-field fill counts before and
    after, and the specific errors (row + identifier + exception). This is
    the machine-readable surface Hermes should read to decide what (if
    anything) needs fixing, rather than parsing stdout. The process also
    exits non-zero if any row errored, even though the CSV/Airtable push
    still completed for every other row.
"""

import argparse
import csv
import functools
import json
import os
import re
import sqlite3
import sys
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    import requests
except ImportError:
    sys.exit("Missing dependency: pip install requests openpyxl --break-system-packages")

try:
    import dns.resolver as _dns_resolver  # optional — only used for the free-scrape MX check
except ImportError:
    _dns_resolver = None

CACHE_DB = "enrich_cache.db"
RATE_LIMIT_SECONDS = {"hunter": 1.0, "apollo": 1.0, "pdl": 1.0, "airtable": 0.25, "scrape": 1.5, "smtp": 2.0}

TARGET_FIELDS = [
    "first_name", "last_name", "business_name", "phone", "email",
    "linkedin", "website", "address", "categories", "description",
    "rating", "review_count", "social_links", "website_score", "source",
    "enrichment_notes",
]


@dataclass
class Lead:
    first_name: str = ""
    last_name: str = ""
    business_name: str = ""
    phone: str = ""
    email: str = ""
    linkedin: str = ""
    website: str = ""
    address: str = ""
    categories: str = ""
    description: str = ""
    rating: str = ""
    review_count: str = ""
    social_links: str = ""
    website_score: str = ""
    source: str = ""
    _enrichment_notes: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Cache
# ---------------------------------------------------------------------------

def init_cache():
    conn = sqlite3.connect(CACHE_DB)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS lookups (
            provider TEXT, lookup_key TEXT, result_json TEXT, fetched_at REAL,
            PRIMARY KEY (provider, lookup_key)
        )
    """)
    conn.commit()
    return conn


def cache_get(conn, provider, key):
    row = conn.execute(
        "SELECT result_json FROM lookups WHERE provider=? AND lookup_key=?",
        (provider, key),
    ).fetchone()
    return json.loads(row[0]) if row else None


def cache_set(conn, provider, key, result):
    conn.execute(
        "INSERT OR REPLACE INTO lookups VALUES (?,?,?,?)",
        (provider, key, json.dumps(result), time.time()),
    )
    conn.commit()


_last_call = {}


def _throttle(provider):
    wait = RATE_LIMIT_SECONDS.get(provider, 1.0)
    last = _last_call.get(provider, 0)
    elapsed = time.time() - last
    if elapsed < wait:
        time.sleep(wait - elapsed)
    _last_call[provider] = time.time()


# ---------------------------------------------------------------------------
# Provider calls — cache-aware, rate-limited, safe to call with --dry-run and
# zero keys set. All go through _request_with_retry so a single flaky response
# from one provider can't take down the whole batch (see main()'s per-row
# try/except for the second half of that guarantee).
# ---------------------------------------------------------------------------

def _request_with_retry(method, url, max_retries=2, backoff=1.0, **kwargs):
    """requests.request wrapper that retries transient failures (timeouts,
    connection errors, 5xx) with exponential backoff. 4xx responses are
    returned as-is, not retried — those are real answers (bad request, not
    found, bad auth), and retrying them just burns quota for the same result."""
    for attempt in range(max_retries + 1):
        try:
            resp = requests.request(method, url, **kwargs)
        except requests.RequestException:
            if attempt < max_retries:
                time.sleep(backoff * (2 ** attempt))
                continue
            raise
        if resp.status_code >= 500 and attempt < max_retries:
            time.sleep(backoff * (2 ** attempt))
            continue
        return resp


def hunter_find_email(conn, domain, first_name, last_name, dry_run=False):
    key = f"{domain}:{first_name}:{last_name}"
    cached = cache_get(conn, "hunter_find", key)
    if cached:
        return cached
    if dry_run:
        return {"email": None, "confidence": None, "dry_run": True}
    api_key = os.environ.get("HUNTER_API_KEY")
    if not api_key:
        raise RuntimeError("HUNTER_API_KEY not set")
    _throttle("hunter")
    resp = _request_with_retry(
        "GET", "https://api.hunter.io/v2/email-finder",
        params={"domain": domain, "first_name": first_name, "last_name": last_name, "api_key": api_key},
        timeout=15,
    )
    resp.raise_for_status()
    data = resp.json().get("data", {})
    result = {"email": data.get("email"), "confidence": data.get("score")}
    cache_set(conn, "hunter_find", key, result)
    return result


def hunter_verify_email(conn, email, dry_run=False):
    cached = cache_get(conn, "hunter_verify", email)
    if cached:
        return cached
    if dry_run:
        return {"status": "dry_run", "score": None}
    api_key = os.environ.get("HUNTER_API_KEY")
    if not api_key:
        raise RuntimeError("HUNTER_API_KEY not set")
    _throttle("hunter")
    resp = _request_with_retry(
        "GET", "https://api.hunter.io/v2/email-verifier",
        params={"email": email, "api_key": api_key},
        timeout=15,
    )
    resp.raise_for_status()
    data = resp.json().get("data", {})
    result = {"status": data.get("status"), "score": data.get("score")}
    cache_set(conn, "hunter_verify", email, result)
    return result


def apollo_enrich_person(conn, first_name, last_name, company, dry_run=False):
    key = f"{first_name}:{last_name}:{company}"
    cached = cache_get(conn, "apollo", key)
    if cached:
        return cached
    if dry_run:
        return {"email": None, "title": None, "linkedin": None, "dry_run": True}
    api_key = os.environ.get("APOLLO_API_KEY")
    if not api_key:
        raise RuntimeError("APOLLO_API_KEY not set")
    _throttle("apollo")
    resp = _request_with_retry(
        "POST", "https://api.apollo.io/v1/people/match",
        headers={"Cache-Control": "no-cache", "Content-Type": "application/json"},
        json={"api_key": api_key, "first_name": first_name, "last_name": last_name, "organization_name": company},
        timeout=15,
    )
    resp.raise_for_status()
    person = resp.json().get("person") or {}
    result = {
        "email": person.get("email"),
        "title": person.get("title"),
        "linkedin": person.get("linkedin_url"),
    }
    cache_set(conn, "apollo", key, result)
    return result


def pdl_enrich_phone(conn, phone, dry_run=False):
    cached = cache_get(conn, "pdl", phone)
    if cached:
        return cached
    if dry_run:
        return {"full_name": None, "company": None, "linkedin": None, "dry_run": True}
    api_key = os.environ.get("PDL_API_KEY")
    if not api_key:
        raise RuntimeError("PDL_API_KEY not set")
    _throttle("pdl")
    resp = _request_with_retry(
        "GET", "https://api.peopledatalabs.com/v5/person/enrich",
        headers={"X-Api-Key": api_key},
        params={"phone": phone},
        timeout=15,
    )
    data = resp.json().get("data", {}) if resp.ok else {}
    result = {
        "full_name": data.get("full_name"),
        "company": data.get("job_company_name"),
        "linkedin": data.get("linkedin_url"),
    }
    cache_set(conn, "pdl", phone, result)
    return result


# ---------------------------------------------------------------------------
# Free scrape fallback — used by enrich_domain_based ONLY when HUNTER_API_KEY
# isn't set. Scoped to the lead's own domain (its homepage / contact / about
# pages), never a third party — that's a different thing from the LinkedIn /
# search-engine scraping already ruled out per the brief (see module docstring).
# No API key required, so this stays safe to call with zero keys configured.
# ---------------------------------------------------------------------------

_SCRAPE_PATHS = ("", "contact", "contact-us", "about", "about-us")
_SCRAPE_HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; TactikLeadEnrichment/1.0; +free-scrape-fallback)"}
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
_JUNK_DOMAINS = {"example.com", "sentry.io", "wixpress.com", "godaddy.com", "schema.org",
                  "w3.org", "cloudflare.com", "wordpress.com", "gravatar.com"}
_JUNK_LOCALPARTS = {"noreply", "no-reply", "donotreply", "do-not-reply", "webmaster",
                     "postmaster", "yourname", "example", "test"}
# Some leads' `website` field is actually a social/profile-page link, not their own
# site. Scraping those isn't "the lead's own site" any more — it's the platform —
# and edges toward the search-engine/social scraping already ruled out per the
# brief, so these are skipped outright rather than fetched.
_SOCIAL_PLATFORM_DOMAINS = {
    "facebook.com", "instagram.com", "linkedin.com", "twitter.com", "x.com",
    "youtube.com", "tiktok.com", "linktr.ee", "wa.me", "m.me", "maps.google.com",
    "g.page", "yelp.com",
}


def _is_social_platform(domain):
    bare = domain.split(":")[0].lower()
    parts = bare.split(".")
    return any(bare == d or bare.endswith("." + d) for d in _SOCIAL_PLATFORM_DOMAINS) or len(parts) < 2


def _scrape_site_emails(domain):
    """Fetch a handful of public pages on the lead's own site and pull out any
    emails found in the raw HTML (mailto: links included, via the same regex)."""
    if _is_social_platform(domain):
        return set()
    found = set()
    for path in _SCRAPE_PATHS:
        for scheme in ("https://", "http://"):
            url = f"{scheme}{domain}/{path}"
            try:
                resp = requests.get(url, headers=_SCRAPE_HEADERS, timeout=8)
            except requests.RequestException:
                continue
            if resp.ok:
                found.update(m.lower() for m in EMAIL_RE.findall(resp.text))
                break  # got a response for this path, no need to also try the other scheme
        if found:
            break  # already have candidates, don't keep hammering more pages
    return found


def _pick_best_email(candidates, domain, first_name, last_name):
    bare_domain = domain.split(":")[0].lower()
    on_domain = [
        e for e in candidates
        if e.split("@", 1)[1] == bare_domain
        and e.split("@", 1)[0] not in _JUNK_LOCALPARTS
        and e.split("@", 1)[1] not in _JUNK_DOMAINS
    ]
    if not on_domain:
        return None
    fn, ln = (first_name or "").lower(), (last_name or "").lower()
    for e in on_domain:
        local = e.split("@", 1)[0]
        if (fn and fn in local) or (ln and ln in local):
            return e
    for role in ("info", "contact", "hello", "enquiries", "sales", "admin"):
        for e in on_domain:
            if e.split("@", 1)[0] == role:
                return e
    return sorted(on_domain)[0]


def _domain_has_mx(conn, domain):
    """Best-effort MX check — confirms the domain CAN receive mail. This is not a
    mailbox verification like Hunter's verifier, just a sanity check for the free
    fallback. Returns None (not False) if dnspython isn't installed, so callers can
    tell 'checked and failed' apart from 'couldn't check'. Cached like every other
    lookup here — a domain's MX record doesn't change often enough to justify
    re-querying DNS on every re-run."""
    cached = cache_get(conn, "mx_check", domain)
    if cached is not None:
        return cached["mx_ok"]
    if _dns_resolver is None:
        return None
    try:
        mx_ok = len(_dns_resolver.resolve(domain, "MX", lifetime=5)) > 0
    except Exception:
        mx_ok = False
    cache_set(conn, "mx_check", domain, {"mx_ok": mx_ok})
    return mx_ok


def scrape_find_email(conn, domain, first_name, last_name, dry_run=False):
    key = f"{domain}:{first_name}:{last_name}"
    cached = cache_get(conn, "scrape_find", key)
    if cached:
        return cached
    if dry_run:
        return {"email": None, "confidence": None, "dry_run": True}
    _throttle("scrape")
    candidates = _scrape_site_emails(domain)
    email = _pick_best_email(candidates, domain, first_name, last_name)
    result = {"email": email, "confidence": "scraped" if email else None}
    cache_set(conn, "scrape_find", key, result)
    return result


# ---------------------------------------------------------------------------
# OPT-IN free pattern-guess + SMTP verify — only runs when --smtp-verify is
# passed AND the site scrape above found nothing. This is a materially
# different footprint from the rest of the pipeline: it opens a real SMTP
# connection to the LEAD'S OWN mail server (never a third party) and issues a
# HELO/MAIL FROM/RCPT TO, reading whether the server accepts the address,
# then quits BEFORE the DATA stage — no message is ever sent. Still, it's an
# active probe against someone else's infrastructure, which some mail
# servers/networks treat as abuse, and outbound port 25 is blocked by many
# cloud providers by default (fails fast/cleanly either way — see except
# below). That's why this stays off unless explicitly requested.
# Caveat: a "catch-all" domain (accepts RCPT TO for anything) will make every
# candidate look "accepted" — this is a real false-positive source, distinct
# from Hunter's verifier which tracks catch-all status explicitly. Every hit
# from this path is flagged in _enrichment_notes with that caveat attached.
# ---------------------------------------------------------------------------

_GENERIC_LOCALPARTS = ("info", "contact", "hello")


def _generate_email_candidates(domain, first_name, last_name, limit=6):
    """Most-likely-first list of candidate addresses: person-based patterns
    (if we have a name) followed by common generic business inboxes."""
    bare_domain = domain.split(":")[0].lower()
    fn = re.sub(r"[^a-z]", "", (first_name or "").lower())
    ln = re.sub(r"[^a-z]", "", (last_name or "").lower())
    candidates = []
    if fn and ln:
        candidates += [f"{fn}.{ln}@{bare_domain}", f"{fn[0]}{ln}@{bare_domain}", f"{fn}@{bare_domain}"]
    elif fn:
        candidates.append(f"{fn}@{bare_domain}")
    candidates += [f"{role}@{bare_domain}" for role in _GENERIC_LOCALPARTS]
    seen, out = set(), []
    for c in candidates:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out[:limit]


def _smtp_verify_email(conn, email, timeout=8):
    """Best-effort SMTP-level check via RCPT TO — connects to the domain's MX
    host and asks whether it would accept mail to this address, without
    sending anything (QUIT before DATA). Cached like every other lookup here,
    so a given address is never re-probed across runs. Returns "accepted",
    "rejected", or "unknown" (couldn't determine: no MX, connection refused/
    blocked, timeout, greylisted, etc. — deliberately not treated the same as
    a confirmed rejection)."""
    cached = cache_get(conn, "smtp_verify", email)
    if cached is not None:
        return cached["status"]
    status = "unknown"
    if _dns_resolver is not None:
        domain = email.split("@", 1)[1]
        try:
            import smtplib
            mx_records = sorted(_dns_resolver.resolve(domain, "MX", lifetime=5), key=lambda r: r.preference)
            mx_host = str(mx_records[0].exchange).rstrip(".")
            with smtplib.SMTP(mx_host, 25, timeout=timeout) as smtp:
                smtp.helo("verify.tactik-enrichment.local")
                smtp.mail("verify@tactik-enrichment.local")
                code, _ = smtp.rcpt(email)
                if code == 250:
                    status = "accepted"
                elif code in (550, 551, 553):
                    status = "rejected"
        except Exception:
            status = "unknown"  # connection refused/blocked, timeout, DNS failure, etc.
    cache_set(conn, "smtp_verify", email, {"status": status})
    return status


def pattern_guess_and_verify(conn, domain, first_name, last_name, dry_run=False):
    """Tries each generated candidate in order, stopping at the first one the
    mail server accepts. Returns {"email": ...} or None. No-ops entirely
    under --dry-run or if dnspython isn't installed."""
    if dry_run or _dns_resolver is None:
        return None
    for candidate in _generate_email_candidates(domain, first_name, last_name):
        _throttle("smtp")
        if _smtp_verify_email(conn, candidate) == "accepted":
            return {"email": candidate}
    return None


# ---------------------------------------------------------------------------
# Per-source loaders — field maps verified against the ACTUAL headers in each
# uploaded file (checked directly, not guessed).
# ---------------------------------------------------------------------------

def _clean(v):
    return str(v).strip() if v is not None else ""


def load_fresh_networking(path):
    """fresh_networking_scored.xlsx — real headers:
    Name, Company, Business Types, Hubs, Email (Profile), Emails (Website), Website,
    Website Score, Lead Score, LinkedIn, Facebook, SSL, Description, Profile URL
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _clean(row[idx["Name"]])
        parts = name.split(" ", 1) if name else ["", ""]
        first = parts[0]
        last = parts[1] if len(parts) > 1 else ""
        email = _clean(row[idx["Email (Profile)"]]) or _clean(row[idx.get("Emails (Website)", idx["Email (Profile)"])])
        leads.append(Lead(
            first_name=first, last_name=last,
            business_name=_clean(row[idx["Company"]]),
            email=email,
            linkedin=_clean(row[idx["LinkedIn"]]),
            website=_clean(row[idx["Website"]]),
            categories=_clean(row[idx["Business Types"]]),
            description=_clean(row[idx["Description"]]),
            website_score=_clean(row[idx["Website Score"]]),
            social_links=_clean(row[idx["Facebook"]]),
            source="fresh_networking",
        ))
    return leads


def load_mazenod(path):
    """mazenod_alumni.xlsx — real headers:
    firstName, lastName, ..., cityName, stateName, country, linkedInUrl, facebookUrl, ...
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        city = _clean(row[idx["cityName"]])
        state = _clean(row[idx["stateName"]])
        country = _clean(row[idx["country"]])
        address = ", ".join(p for p in [city, state, country] if p)
        leads.append(Lead(
            first_name=_clean(row[idx["firstName"]]),
            last_name=_clean(row[idx["lastName"]]),
            linkedin=_clean(row[idx["linkedInUrl"]]),
            social_links=_clean(row[idx["facebookUrl"]]),
            address=address,
            source="mazenod",
        ))
    return leads


def load_transcend(path):
    """transcend_final.xlsx — real headers:
    Name, First Name, Location, Emails, Website, LinkedIn, Facebook, Instagram,
    Twitter, Tags, Bio, Member Since, Profile URL
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name = _clean(row[idx["Name"]])
        first = _clean(row[idx["First Name"]]) or (full_name.split(" ", 1)[0] if full_name else "")
        last = full_name[len(first):].strip() if full_name.startswith(first) else ""
        emails = _clean(row[idx["Emails"]])
        first_email = emails.split(",")[0].strip() if emails else ""
        leads.append(Lead(
            first_name=first, last_name=last,
            email=first_email,
            website=_clean(row[idx["Website"]]),
            linkedin=_clean(row[idx["LinkedIn"]]),
            social_links=_clean(row[idx["Facebook"]]),
            categories=_clean(row[idx["Tags"]]),
            description=_clean(row[idx["Bio"]]),
            address=_clean(row[idx["Location"]]),
            source="transcend",
        ))
    return leads


def load_vsbn(path):
    """vsbn_final.xlsx — real headers:
    Business Name, Phone, Website, Emails, Categories, Description, Rating,
    Review Count, Address, Suburb, State, Postcode. No LinkedIn column exists.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        addr_parts = [
            _clean(row[idx["Address"]]), _clean(row[idx["Suburb"]]),
            _clean(row[idx["State"]]), _clean(row[idx["Postcode"]]),
        ]
        leads.append(Lead(
            business_name=_clean(row[idx["Business Name"]]),
            phone=_clean(row[idx["Phone"]]),
            website=_clean(row[idx["Website"]]),
            email=_clean(row[idx["Emails"]]),
            categories=_clean(row[idx["Categories"]]),
            description=_clean(row[idx["Description"]]),
            rating=_clean(row[idx["Rating"]]),
            review_count=_clean(row[idx["Review Count"]]),
            address=", ".join(p for p in addr_parts if p),
            source="vsbn",
        ))
    return leads


def load_bali_group(path):
    """whatsapp_contacts_bali.json — real structure is {"named": [...], "phones_only": [...]},
    NOT a flat list. named entries have name/phone/group; phones_only is raw phone strings.
    """
    with open(path) as f:
        raw = json.load(f)
    leads = []
    for c in raw.get("named", []):
        name = _clean(c.get("name"))
        parts = name.split(" ", 1) if name else ["", ""]
        leads.append(Lead(
            first_name=parts[0],
            last_name=parts[1] if len(parts) > 1 else "",
            phone=_clean(c.get("phone")).rstrip("-"),
            categories=_clean(c.get("group")),
            source="bali_group",
        ))
    for phone in raw.get("phones_only", []):
        leads.append(Lead(
            phone=_clean(phone).rstrip("-"),
            source="bali_group",
            _enrichment_notes=["no name — phone-only contact"],
        ))
    return leads


SOURCE_LOADERS = {
    "fresh": load_fresh_networking,
    "mazenod": load_mazenod,
    "transcend": load_transcend,
    "vsbn": load_vsbn,
    "bali": load_bali_group,
}


# ---------------------------------------------------------------------------
# Enrichment strategies — one per source
# ---------------------------------------------------------------------------

def enrich_fresh(conn, lead: Lead, dry_run):
    # Already ~98% has an email. Just normalize — no paid API calls needed by default.
    if not lead.email:
        lead._enrichment_notes.append("missing email despite Fresh Networking's high coverage — manual review")
    return lead


def enrich_mazenod(conn, lead: Lead, dry_run):
    lead._enrichment_notes.append(
        "no viable enrichment anchor at scale (0% emails, 4% LinkedIn, no employer field) — "
        "confirmed not enrichable via API; do not spend API budget here without a different data strategy"
    )
    return lead


def enrich_domain_based(conn, lead: Lead, dry_run, smtp_verify=False):
    """Shared strategy for VSBN / Transcend: has a website, missing email."""
    if lead.email or not lead.website:
        return lead
    domain = lead.website.replace("https://", "").replace("http://", "").split("/")[0]

    if dry_run or os.environ.get("HUNTER_API_KEY"):
        guess = hunter_find_email(conn, domain, lead.first_name or "info", lead.last_name or "", dry_run)
        if guess.get("email"):
            verify = hunter_verify_email(conn, guess["email"], dry_run)
            if dry_run or verify.get("status") in ("valid", "accept_all"):
                lead.email = guess["email"]
            else:
                lead._enrichment_notes.append(f"email found but failed verification: {guess['email']} ({verify.get('status')})")
    else:
        # No HUNTER_API_KEY set — free fallback: scrape the LEAD'S OWN site instead
        # of calling Hunter. Lower confidence: no pattern-matching, MX-checked only
        # (not a real mailbox verification), so every hit here is flagged in
        # _enrichment_notes rather than trusted the same as a Hunter result.
        guess = scrape_find_email(conn, domain, lead.first_name, lead.last_name, dry_run)
        if guess.get("email"):
            lead.email = guess["email"]
            mx_ok = _domain_has_mx(conn, guess["email"].split("@", 1)[1])
            note = "email found via free site scrape, not Hunter-verified"
            if mx_ok is False:
                note += " — domain has no MX record, likely undeliverable"
            elif mx_ok is None:
                note += " — MX check unavailable (dnspython not installed)"
            else:
                note += " — domain has a valid MX record (mail-capable, not a full mailbox verification)"
            lead._enrichment_notes.append(note)
        elif smtp_verify and not _is_social_platform(domain):
            # Opt-in only (--smtp-verify): the scrape found nothing published,
            # so try pattern-guessed addresses and confirm each via a real
            # (no-send) SMTP RCPT TO against the lead's own mail server.
            guess2 = pattern_guess_and_verify(conn, domain, lead.first_name, lead.last_name, dry_run)
            if guess2 and guess2.get("email"):
                lead.email = guess2["email"]
                lead._enrichment_notes.append(
                    "email found via free pattern-guess + SMTP verification (opt-in --smtp-verify), "
                    "not Hunter-verified — mail server accepted RCPT TO for this address; "
                    "catch-all domains can produce false positives here"
                )
    return lead


def enrich_bali(conn, lead: Lead, dry_run):
    if lead.phone:
        pdl = pdl_enrich_phone(conn, lead.phone, dry_run)
        if pdl.get("company"):
            lead.business_name = pdl["company"]
        if pdl.get("linkedin"):
            lead.linkedin = pdl["linkedin"]
        if not lead.first_name and pdl.get("full_name"):
            parts = pdl["full_name"].split(" ")
            lead.first_name, lead.last_name = parts[0], " ".join(parts[1:])
    if lead.business_name and lead.first_name:
        apollo = apollo_enrich_person(conn, lead.first_name, lead.last_name, lead.business_name, dry_run)
        if apollo.get("email"):
            lead.email = apollo["email"]
        if apollo.get("linkedin") and not lead.linkedin:
            lead.linkedin = apollo["linkedin"]
    return lead


ENRICH_STRATEGIES = {
    "fresh": enrich_fresh,
    "mazenod": enrich_mazenod,
    "vsbn": enrich_domain_based,
    "transcend": enrich_domain_based,
    "bali": enrich_bali,
}


# ---------------------------------------------------------------------------
# Airtable push — final step, optional, does not gate the CSV output
# ---------------------------------------------------------------------------

def airtable_push(rows, table_name, dry_run=False):
    api_key = os.environ.get("AIRTABLE_API_KEY")
    base_id = os.environ.get("AIRTABLE_BASE_ID")
    if not api_key or not base_id:
        raise RuntimeError("AIRTABLE_API_KEY / AIRTABLE_BASE_ID not set")

    url = f"https://api.airtable.com/v0/{base_id}/{table_name}"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    pushed, failed = 0, 0
    # Airtable's create endpoint accepts up to 10 records per request.
    for i in range(0, len(rows), 10):
        batch = rows[i:i + 10]
        payload = {"records": [{"fields": r} for r in batch]}
        if dry_run:
            pushed += len(batch)
            continue
        _throttle("airtable")
        try:
            resp = _request_with_retry("POST", url, headers=headers, json=payload, timeout=15)
        except requests.RequestException as exc:
            failed += len(batch)
            print(f"  Airtable batch failed (network error, not retried further): {exc}", file=sys.stderr)
            continue
        if resp.ok:
            pushed += len(batch)
        else:
            failed += len(batch)
            print(f"  Airtable batch failed ({resp.status_code}): {resp.text[:200]}", file=sys.stderr)
    return pushed, failed


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _target_snapshot(lead):
    """The real enrichment fields only (excludes enrichment_notes, which isn't
    a dataclass field — see _output_row) — used to decide whether a row was
    really enriched, as opposed to just picking up an _enrichment_notes entry
    (e.g. mazenod's blanket "not enrichable" note, or an error note), which
    isn't the same thing. Also the basis for the run summary's fill counts."""
    return {k: v for k, v in asdict(lead).items() if k in TARGET_FIELDS}


def _output_row(lead):
    """What actually gets written to the CSV/Airtable: the target fields plus
    a human-readable enrichment_notes column, so a row flagged for manual
    review (free-scrape guess, failed verification, enrichment error, ...) is
    visible as exactly that in Airtable — not only in the run's summary.json."""
    row = _target_snapshot(lead)
    row["enrichment_notes"] = "; ".join(lead._enrichment_notes)
    return row


def _field_fill_counts(leads):
    """Per-field non-empty counts across a batch — used to report real
    before/after coverage in the run summary, not just a row-count guess."""
    counts = {f: 0 for f in TARGET_FIELDS}
    for lead in leads:
        snap = _target_snapshot(lead)
        for f in TARGET_FIELDS:
            if snap.get(f):
                counts[f] += 1
    return counts


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--source", required=True, choices=list(SOURCE_LOADERS.keys()))
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--output", required=True, type=Path)
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--airtable-push", action="store_true")
    ap.add_argument("--airtable-table", type=str, default=None)
    ap.add_argument(
        "--smtp-verify", action="store_true",
        help="OPT-IN, vsbn/transcend only, no-op unless HUNTER_API_KEY is unset. When the free "
             "site-scrape fallback finds nothing, also try pattern-guessed addresses "
             "(first.last@domain, info@domain, etc.) confirmed via a real (no-send) SMTP "
             "RCPT TO against the lead's own mail server. Off by default: this opens outbound "
             "SMTP connections to third-party servers, which some networks treat as probing/abuse.",
    )
    args = ap.parse_args()

    if not args.input.exists():
        sys.exit(f"Input file not found: {args.input}")
    if args.airtable_push and not args.airtable_table:
        sys.exit("--airtable-table is required when using --airtable-push")

    started_at = datetime.now(timezone.utc)
    conn = init_cache()
    leads = SOURCE_LOADERS[args.source](args.input)
    if args.limit:
        leads = leads[: args.limit]

    field_fill_before = _field_fill_counts(leads)

    # Per-row try/except so one bad API response, timeout, or malformed row
    # can't take down an entire batch — earlier versions let a raised
    # exception from e.g. hunter_find_email kill the whole run mid-way,
    # discarding hours of already-completed (but not-yet-written) enrichment.
    # The CSV is now also written incrementally (flushed every row), so a
    # hard kill (OOM, container restart) leaves a usable partial file instead
    # of nothing — re-running picks up instantly for already-cached lookups.
    strategy = ENRICH_STRATEGIES[args.source]
    if strategy is enrich_domain_based:
        strategy = functools.partial(strategy, smtp_verify=args.smtp_verify)
    enriched, notes, errored = 0, 0, 0
    errors = []

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=TARGET_FIELDS)
        writer.writeheader()
        for i, lead in enumerate(leads):
            before = _target_snapshot(lead)
            try:
                strategy(conn, lead, args.dry_run)
            except Exception as exc:
                errored += 1
                identifier = (
                    lead.email or lead.business_name or lead.phone
                    or f"{lead.first_name} {lead.last_name}".strip() or f"row {i}"
                )
                errors.append({"row_index": i, "identifier": identifier, "error": f"{type(exc).__name__}: {exc}"})
                lead._enrichment_notes.append(f"enrichment error, row left as-is: {type(exc).__name__}: {exc}")
                print(f"  [{i}] enrichment error, continuing: {type(exc).__name__}: {exc}", file=sys.stderr)
            if _target_snapshot(lead) != before:
                enriched += 1
            if lead._enrichment_notes:
                notes += 1
            writer.writerow(_output_row(lead))
            f.flush()

    field_fill_after = _field_fill_counts(leads)
    rows_for_output = [_output_row(lead) for lead in leads]

    print(
        f"{args.source}: {len(leads)} rows processed, {enriched} enriched, "
        f"{notes} flagged for manual review" + (f", {errored} errored" if errored else "")
    )
    print(f"-> {args.output}")
    if args.dry_run:
        print("(dry run — no paid API calls were made)")

    airtable_result = None
    if args.airtable_push:
        pushed, failed = airtable_push(rows_for_output, args.airtable_table, args.dry_run)
        airtable_result = {"attempted": True, "pushed": pushed, "failed": failed}
        print(f"Airtable push: {pushed} pushed, {failed} failed" + (" (dry run)" if args.dry_run else ""))

    # Machine-readable run summary — this is what Hermes should parse to decide
    # whether/what to fix, rather than scraping stdout. Sits next to the CSV as
    # <output>.summary.json (e.g. vsbn_enriched.csv -> vsbn_enriched.summary.json).
    summary = {
        "source": args.source,
        "input": str(args.input),
        "output": str(args.output),
        "started_at": started_at.isoformat(),
        "finished_at": datetime.now(timezone.utc).isoformat(),
        "dry_run": args.dry_run,
        "rows_total": len(leads),
        "rows_enriched": enriched,
        "rows_flagged_for_review": notes,
        "rows_errored": errored,
        "field_fill_before": field_fill_before,
        "field_fill_after": field_fill_after,
        "errors": errors,
        "airtable": airtable_result,
    }
    summary_path = args.output.with_suffix(".summary.json")
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"-> {summary_path} (machine-readable run summary for Hermes)")

    # Non-zero exit on partial failure — the CSV/Airtable push still completed
    # for every other row, but an orchestrator (Hermes) should know something
    # needs attention rather than silently treating this as a clean run.
    if errored:
        sys.exit(1)


if __name__ == "__main__":
    main()
